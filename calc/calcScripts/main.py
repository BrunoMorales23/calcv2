import os
from .initialize import *
from .settings import *
from .initialize.queue import *
from .log import *
from .llama import *
import sys
from openpyxl import load_workbook
import re
from openpyxl.styles import Font, PatternFill, Alignment
from main.models import workQueue
from django.core.serializers.json import DjangoJSONEncoder
import json

def initializeScripts():
    sys.modules['tarfile'] = None
    sys.modules['pickle'] = None
    settings.init()
    #work_queue = queue.Queue()
    #log_path = setLog()
    work_queue = []

    return work_queue #log_path

def wqUpload(work_queue, file_name, file_path):

    print(f"Archivo actual: {settings.inputPath}{file_name}")
    queue_id = file_name.replace(".pdf","")
    #work_queue.enqueue(Node(path=file_path ,id=queue_id))
    workQueue.objects.create(id_value=queue_id, path_value=file_path)
    new_item = {"path":file_path ,"id":queue_id}
    work_queue.append(new_item)
    print(f"Item con ID: {queue_id} cargado en cola.")

    return work_queue

#Step 1
def ocrExecution(item_id, item_path):

    item_content = pdfToText(settings.tesseractPath, settings.popplerPath, item_path)

    print("--- OCR Run Finished ---")
    workQueue.objects.filter(id_value=item_id, path_value=item_path, status="Pendiente").update(log="Inicializando Ejecución",status="Ejecutando")

    return item_content, 2

#Step 2
def llamaExecution(item_id, item_path, item_content):
    ollama = llama.Llama(content=item_content)
    llama_template = ollama.getTemplate(settings.promptBase)
    llama_result = ollama.executePrompt(str(llama_template), item_content)

    print(f"{item_id}--- Llama Run Finished ---")
    workQueue.objects.filter(id_value=item_id, path_value=item_path, status="Ejecutando").update(log="Finalizado el Análisis por IA")

    try:
        if os.path.exists(item_path):
            os.remove(item_path)
    except Exception as e:
        print(f"No se pudo eliminar el documento con path: {item_path}. Excepción: {e}")
        pass
    del ollama
    return llama_result, 3

#Step 3
def writeOutput(llama_result, item_id, item_path):

    current_row = 3
    current_value = 0
    workBook = load_workbook(settings.template_xlsx_output)
    workSheet = workBook.active

    splittedLines = [p.strip() for p in llama_result.split("Paso") if p.strip()]
    regex_tittle = r"\]: (.*?) . Estimación:"
    regex_horas = r"Estimación:\s*(.*?)(?:\s*Paso|$)"

    for line in splittedLines:
        task_tittle = re.findall(regex_tittle, line)
        if task_tittle == None or task_tittle == "":
            regex_tittle = r"\]:\s*(.*?)\s*\. Estimación:"
            task_tittle = re.findall(regex_tittle, line)
        task_time = re.findall(regex_horas, line)
        print(task_tittle)

        fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        cell_b = workSheet[f"B{current_row}"]
        cell_c = workSheet[f"C{current_row}"]

        cell_b.value = task_tittle[0] if task_tittle else ""
        cell_c.value = task_time[0] if task_time else ""

        cell_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_c.alignment = Alignment(horizontal='center', vertical='center')
        cell_b.font = Font(name='Calibri', size=11, color="000000")
        cell_c.font = Font(name='Calibri', size=11, color="000000")

        if current_row % 2 == 0:
            cell_b.fill = fill
            cell_c.fill = fill

        current_row += 1
        current_value += 1

    workBook.save(f"Estimación PDD {item_id}.xlsx")
    workQueue.objects.filter(id_value=item_id, path_value=item_path, status="Ejecutando").update(log="Exportación de datos Completada", status="Completado")
    return 0

def getCurrentItem(log, status):
    return json.loads(json.dumps(list(workQueue.objects.filter(log=log, status=status).values("id_value", "status", "log", "updated_at").order_by("-updated_at")),cls=DjangoJSONEncoder))

#Se separan estas funciones a fin de poder ser utilizadas en casos X desde views.py
# def setLog():
#     current_log_path = initialize.createDir(settings.logsPath)
#     current_log = logs.createNewLog(current_log_path)
#     return current_log

# def writeLog(current_log, message):
#     logs.writeLogValue(current_log, message)

#Función para multi prompt
# def setEstimation(pathPrompt, content):
#     if pathPrompt == "1":
#         pathPrompt = settings.estimation_prompt
#     elif pathPrompt == "2":
#         pathPrompt = settings.horas_prompt
#     ollama = llama.Llama(content=content)
#     with open(pathPrompt, "r", encoding="utf-8") as f:
#         queryValue = f.read()
#         ollama.modifyQuery(queryValue)
#     llama_template = ollama.getTemplate(settings.estimation_prompt)
#     llama_result = ollama.executePrompt(str(llama_template) + queryValue, content)
#     del ollama

#     return llama_result